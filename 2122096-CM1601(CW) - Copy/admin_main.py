import enter_car_details
import input_car_details
import openpyxl

# loading CarDealership.xlsx
workbook = openpyxl.load_workbook("CarDealership.xlsx")

# Declare variables
status = ""
number_list = []
make_list = []
model_list = []
year_list = []
mileage_list = []
owner_info_list = []
ask_price_list = []
sell_price_list = []
car_status_list = []
last_price_status_list = []
last_price_list = []

# this will enable admin to enter car details to the CarDealership.xlsx until he selects exit option
while status != "YES":
    status = ""
    month = input_car_details.type_month()
    make = input_car_details.type_make()
    model = input_car_details.type_model()
    year = input_car_details.type_year()
    mileage = input_car_details.type_mileage()
    owner_info = input_car_details.type_owner_info()
    ask_price = input_car_details.type_ask_price()
    sell_price = input_car_details.type_sell_price()
    car_status = input_car_details.type_car_status()
    last_price_status = input_car_details.type_last_price_status()
    # this is to define the last selling price when the car is under negotiations
    if last_price_status == 'not_fixed':
        last_price = input_car_details.type_last_price()
    else:
        last_price = sell_price

   # all admin entered car details will append to lists
    make_list.append(make)
    model_list.append(model)
    year_list.append(year)
    mileage_list.append(mileage)
    owner_info_list.append(owner_info)
    ask_price_list.append(ask_price)
    sell_price_list.append(sell_price)
    car_status_list.append(car_status)
    last_price_status_list.append(last_price_status)
    last_price_list.append(last_price)

    # to refer to the related sheet according to the admin entered month
    sheet = input_car_details.refer_month(month,workbook)

    max_row = sheet.max_row
    number = sheet.max_row
    number_list.append(number)
    # to enter the car details on CarDealership.xlsx
    enter_car_details.write_record_number(max_row,workbook,sheet,number_list,make_list,model_list,year_list,mileage_list,owner_info_list,ask_price_list,sell_price_list,car_status_list,last_price_status_list,last_price_list)

    # to clear the all list to run the same code again purpose
    number_list.clear()
    make_list.clear()
    model_list.clear()
    year_list.clear()
    mileage_list.clear()
    owner_info_list.clear()
    ask_price_list.clear()
    sell_price_list.clear()
    car_status_list.clear()
    last_price_status_list.clear()
    last_price_list.clear()

    while not (status == 'YES' or status == 'NO'):
        try:
            status = input("Do you want to exit ('YES'/'NO'): ").upper()
            if not (status == 'YES' or status == 'NO'):
                raise Exception("Enter only 'YES' or 'NO'")
        except ValueError as e:
            print(e)
        except Exception as e:
            print(e)






