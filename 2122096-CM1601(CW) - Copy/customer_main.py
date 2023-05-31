import openpyxl
import sort
import view_appointment
import maintenance

# loading CarDealership.xlsx
workbook = openpyxl.load_workbook("CarDealership.xlsx")
# this uses the current records that i have used the current month as march
sheet = workbook['2022MARCH']

# loading CarAppointment.xlsx
workbook_2 = openpyxl.load_workbook("CarAppointment.xlsx")
# this use the current sheet to take data from that
sheet_2 = workbook_2['appointment']

# loading CustomerInformation.xlsx
workbook_3 = openpyxl.load_workbook("CustomerInformation.xlsx")
sheet_3 = workbook_3['2022MARCH']
max_row_3 = sheet_3.max_row

# loading CompanyInformation.xlsx
workbook_4 = openpyxl.load_workbook('CompanyInformation.xlsx')
sheet_4 = workbook_4['2022MARCH']
max_row_4 = sheet_4.max_row

# loading MaintenanceTime.xlsx
workbook_5 = openpyxl.load_workbook('MaintenanceTime.xlsx')
# this uses current sheet to take data from that
sheet_5 = workbook_5['Sheet1']
max_row_5 = sheet_5.max_row

# Declare variables
number_list = []
make_list = []
model_list = []
mileage_list = []
year_list = []
owner_info_list = []
sell_price_list = []
row = ''
appointment = ''
buy_or_not = ''
new_list_1 = []
new_list_2 = []
new_list_3 = []
option = 0
option_1 = 0
car_record = []

max_row = sheet.max_row
max_row_2 = sheet_2.max_row
max_column = sheet_2.max_column

# to print the every available cars
sort.without_any_sort(new_list_2,new_list_3,car_record,option,option_1,max_row,sheet,number_list,make_list,model_list,year_list,mileage_list,owner_info_list,sell_price_list)

number_list.clear()
make_list.clear()
model_list.clear()
mileage_list.clear()
year_list.clear()
owner_info_list.clear()
sell_price_list.clear()
new_list_1.clear()
new_list_2.clear()
new_list_3.clear()
car_record.clear()


print("""If you want to sort the above details\nby mileage --> press 1\nby budget --> press 2
           \nby budget and mileage --> press 3\nif not --> press 4""")


option = sort.enter_first_option()


if option == 1:
    user_prefer_mileage = sort.enter_user_prefer_mileage()
    option_1 = sort.enter_second_option()
    sort.sort_by_mileage(new_list_2,new_list_3,car_record,option,option_1,user_prefer_mileage,max_row,sheet,number_list,make_list,model_list,year_list,mileage_list,owner_info_list,sell_price_list)

elif option == 2:
    budget = sort.enter_user_budget()
    option_1 = sort.enter_second_option()
    sort.sort_by_budget(new_list_2,new_list_3,car_record,option,option_1,budget,max_row,sheet,number_list,make_list,model_list,year_list,mileage_list,owner_info_list,sell_price_list)

elif option == 3:
    user_prefer_mileage = sort.enter_user_prefer_mileage()
    budget = sort.enter_user_budget()
    option = sort.enter_third_option()
    option_1 = sort.enter_second_option()
    sort.sort_by_mileage_and_budget(new_list_2,new_list_3,car_record,option,option_1,budget,user_prefer_mileage,max_row,sheet,number_list,make_list,model_list,year_list,mileage_list,owner_info_list,sell_price_list)


print("""If you want an appointment for a car model ---> press 1\nif not ---> press 2
                           """)
option_1 = view_appointment.enter_first_option()

if option_1 == 1:
    car_num = view_appointment.enter_car_number(max_row,sheet)
    view_appointment.print_appointments(max_row_2, sheet_2, car_num, max_column)

print('If you want to look for a car maintenance --> press 1\nif not --> press 2')
option_1 = view_appointment.enter_first_option()
if option_1 == 1:
    print('If you have purchased a car from me and if you want to look for a maintenance for that --   . press 1\nif not --> press 2')
    option_1 = view_appointment.enter_first_option()
    if option_1 == 1:
        custom_receipt_num = maintenance.define_customer_receipt_num(sheet_3, max_row_3)
        company_receipt_num = maintenance.define_company_receipt_num(sheet_4, max_row_4)
        receipt_num = maintenance.enter_receipt_num(custom_receipt_num, company_receipt_num)
        custom_name = maintenance.define_customer_name(sheet_3, max_row_3)
        company_name = maintenance.define_company_name(sheet_4, max_row)
        name = maintenance.enter_customer_or_company_name(receipt_num,custom_receipt_num,company_receipt_num,custom_name,company_name)
        print('You can service your car make under 10% discount')
        maintenance.print_maintenance_time_1(sheet_5, max_row_5)
    else:
        maintenance.print_maintenance_time_2(sheet_5, max_row_5)
















