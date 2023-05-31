import update_record
import input_car_details
import openpyxl

# load CarDealership.xlsx workbook
workbook = openpyxl.load_workbook("CarDealership.xlsx")

# Declare variables
option_2 = 0
option_3 = 0
number_list = []
make_list = []
model_list = []
year_list = []
mileage_list = []
owner_info_list = []
ask_price_list = []
sell_price_list = []
car_status_list = []
last_price_status_list= []
last_price_list = []
new_list_2 = []
new_list_3 = []
car_record = []

print('If you want to update records ---> press 1\nIf not ---> press 2')
option = update_record.enter_option()
# this will run until admin selects exit option
while option == 1:
    if (option_2 == 0 or option_2 == 3):
        month = input_car_details.type_month()
        sheet = input_car_details.refer_month(month, workbook)
        max_row = sheet.max_row
        update_record.view_details(new_list_2, new_list_3, car_record, max_row, sheet, number_list, make_list,
                                   model_list,
                                   year_list, mileage_list, owner_info_list, ask_price_list, car_status_list,
                                   sell_price_list,last_price_status_list,last_price_list)
        new_list_2.clear()
        new_list_3.clear()
        car_record.clear()
        number_list.clear()
        make_list.clear()
        model_list.clear()
        year_list.clear()
        mileage_list.clear()
        owner_info_list.clear()
        ask_price_list.clear()
        car_status_list.clear()
        sell_price_list.clear()
        last_price_status_list.clear()
        last_price_list.clear()

    if (option_2 == 0 or option_2 == 2 or option_2 == 3 ):
        detail_row_number = update_record.enter_record_number(max_row)
    print('''\ncolumn numbers:\nmake --> column 2\nmodel --> column 3\nyear --> column 4\nmileage --> column 5
            \nowner_info --> column 6\nask_price --> column 7\nprice --> column 8\ncar_status --> column 9\n
            last_price_status(is it contained negotiations or not) --> column 10\nlast_price --> column 11
                  ''')
    column_number = update_record.enter_column_number()
    # to update cells according to admin entered record number and column number
    if column_number == 2:
        make = input_car_details.type_make()
        sheet.cell(row=detail_row_number + 1, column=2, value=make)
        workbook.save('CarDealership.xlsx')
    if column_number == 3:
        model = input_car_details.type_model()
        sheet.cell(row=detail_row_number + 1, column=3, value=model)
        workbook.save('CarDealership.xlsx')
    if column_number == 4:
        year = input_car_details.type_year()
        sheet.cell(row=detail_row_number + 1, column=4, value=year)
        workbook.save('CarDealership.xlsx')

    if column_number == 5:
        mileage = input_car_details.type_mileage()
        sheet.cell(row=detail_row_number + 1, column=5, value=mileage)
        workbook.save('CarDealership.xlsx')

    if column_number == 6:
        owner_info = input_car_details.type_owner_info()
        sheet.cell(row=detail_row_number + 1, column=6, value=owner_info)
        workbook.save('CarDealership.xlsx')
    if column_number == 7:
        ask_price = input_car_details.type_ask_price()
        sheet.cell(row=detail_row_number + 1, column=7, value=ask_price)
        workbook.save('CarDealership.xlsx')

    if column_number == 8:
        sell_price = input_car_details.type_sell_price()
        sheet.cell(row=detail_row_number + 1, column=8, value=sell_price)
        workbook.save('CarDealership.xlsx')
    if column_number == 9:
        car_status = input_car_details.type_car_status()
        sheet.cell(row=detail_row_number + 1, column=9, value=car_status)
        workbook.save('CarDealership.xlsx')

    if column_number == 10:
        last_price_status = input_car_details.type_last_price_status()
        sheet.cell(row=detail_row_number + 1, column=10, value=last_price_status)
        workbook.save('CarDealership.xlsx')

    if column_number == 11:
        last_price = input_car_details.type_last_price()
        sheet.cell(row=detail_row_number + 1, column=11, value=last_price)
        workbook.save('CarDealership.xlsx')



    print('\nIf you want to update records again ---> press 1\nIf not ---> press 2')
    option = update_record.enter_option()
    if option == 2:
        break
    print('''\nIf you want to update the record:\nin same month same row --> press 1
          same month different row --> press 2
          \ndifferent month --> press 3''')
    option_2 = update_record.enter_second_option()
    column_number = ''


print('If you want to delete records ---> press 1\nIf not ---> press 2')
option = update_record.enter_option()
# this will run until admin selects exit option
while option == 1:
    if (option_3 == 0 or option_3 == 2):
        month = input_car_details.type_month()
        sheet = input_car_details.refer_month(month, workbook)
        max_row = sheet.max_row

    if option_3 == 1:
        max_row = sheet.max_row
    update_record.view_details(new_list_2, new_list_3, car_record, max_row, sheet, number_list, make_list,
                               model_list,
                               year_list, mileage_list, owner_info_list, ask_price_list, car_status_list,
                               sell_price_list,last_price_status_list,last_price_list)
    new_list_2.clear()
    new_list_3.clear()
    car_record.clear()
    number_list.clear()
    make_list.clear()
    model_list.clear()
    year_list.clear()
    mileage_list.clear()
    owner_info_list.clear()
    ask_price_list.clear()
    car_status_list.clear()
    sell_price_list.clear()
    last_price_status_list.clear()
    last_price_list.clear()



    detail_row_number = update_record.enter_record_number_to_delete(max_row)
    # to delete the admin entered row number record
    sheet.delete_rows(idx = detail_row_number+1)
    workbook.save('CarDealership.xlsx')

    # after deleting a record this will make the car reference number into correct order
    for i in range((detail_row_number+1),max_row):
        sheet.cell(row = i,column = 1, value = i -1)
        workbook.save('CarDealership.xlsx')
    print('if you want to delete records again ---> press 1\nif not --> press 2')
    option = update_record.enter_option()
    if option == 2:
        break
    print('if you want to delete records in same month ---> press 1\nin different month ---> press 2')
    option_3 = update_record.enter_third_option()