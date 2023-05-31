import best_selling_models
import monthly_profit
import best_customers_or_companies
import sold_details_append
import generate_graph
import maintenance_func
import openpyxl
workbook = openpyxl.load_workbook("CarDealership.xlsx")
workbook_3 = openpyxl.load_workbook("CustomerInformation.xlsx")
workbook_4 = openpyxl.load_workbook("CompanyInformation.xlsx")
workbook_5 = openpyxl.load_workbook('MaintenanceDetails.xlsx')

# to generate report and graph
month = 9
sheet = best_selling_models.refer_month(month,workbook)
max_row = sheet.max_row
model_list = best_selling_models.append_models_to_a_list(max_row,sheet)
model_dict = best_selling_models.make_a_dic_by_count_of_the_models(model_list)
print('Best five selling models for september 2021')
best_selling_models.print_best_five_selling_models_of_a_month(model_dict)
ask_price_list = monthly_profit.create_ask_price_list(sheet,max_row)
sell_price_list = monthly_profit.create_sell_price_list(sheet,max_row)
total_ask_price = monthly_profit.add_ask_prices(ask_price_list)
total_sell_price = monthly_profit.add_sell_prices(sell_price_list)
print('\nThe monthly profit generated by the sales department during september 2021')
monthly_profit_value = monthly_profit.print_monthly_profit(total_sell_price,total_ask_price)
sheet_3 = sold_details_append.refer_month(month,workbook_3)
sheet_4 = sold_details_append.refer_month_in_CompnyDetails_workbook(month,workbook_4)
customer_name = best_customers_or_companies.create_customer_list(sheet_3)
company_name = best_customers_or_companies.create_company_list(sheet_4)
customer_dict = best_customers_or_companies.make_a_dic_by_count_of_the_customers(customer_name)
company_dict = best_customers_or_companies.make_a_dic_by_count_of_the_companies(company_name)
print('\nAll best customers during september 2021')
best_customers_or_companies.find_best_customers(customer_dict)
print('\nAll best companies during september 2021')
best_customers_or_companies.find_best_companies(company_dict)
num_of_customers = sheet_3.max_row - 1
number_of_companies = sheet_4.max_row - 1
number_of_sold_cars = num_of_customers+number_of_companies
num_of_sold_car_list = generate_graph.create_a_list_with_number_of_sold_cars(number_of_sold_cars)
total_income_sell_car = generate_graph.find_total_income_by_selling_car(sheet_3,sheet_4)
sheet_5 = maintenance_func.refer_month(month,workbook_5)
total_maintenance_cost = generate_graph.find_total_income_by_maintenance(sheet_5)
total_income = total_income_sell_car + total_maintenance_cost
total_income_list = generate_graph.create_total_income_list(total_income)

month = 10
sheet = best_selling_models.refer_month(month,workbook)
max_row = sheet.max_row
model_list = best_selling_models.append_models_to_a_list(max_row,sheet)
model_dict = best_selling_models.make_a_dic_by_count_of_the_models(model_list)
print('\nBest five selling models for october 2021')
best_selling_models.print_best_five_selling_models_of_a_month(model_dict)
ask_price_list = monthly_profit.create_ask_price_list(sheet,max_row)
sell_price_list = monthly_profit.create_sell_price_list(sheet,max_row)
total_ask_price = monthly_profit.add_ask_prices(ask_price_list)
total_sell_price = monthly_profit.add_sell_prices(sell_price_list)
print('\nThe monthly profit generated by the sales department during october 2021')
monthly_profit_value = monthly_profit.print_monthly_profit(total_sell_price,total_ask_price)
sheet_3 = sold_details_append.refer_month(month,workbook_3)
sheet_4 = sold_details_append.refer_month_in_CompnyDetails_workbook(month,workbook_4)
customer_name = best_customers_or_companies.create_customer_list(sheet_3)
company_name = best_customers_or_companies.create_company_list(sheet_4)
customer_dict = best_customers_or_companies.make_a_dic_by_count_of_the_customers(customer_name)
company_dict = best_customers_or_companies.make_a_dic_by_count_of_the_companies(company_name)
print('\nAll best customers during october 2021')
best_customers_or_companies.find_best_customers(customer_dict)
print('\nAll best companies during october 2021')
best_customers_or_companies.find_best_companies(company_dict)
num_of_customers = sheet_3.max_row - 1
number_of_companies = sheet_4.max_row - 1
number_of_sold_cars = num_of_customers+number_of_companies
num_of_sold_car_list = generate_graph.create_a_list_with_number_of_sold_cars(number_of_sold_cars)
total_income_sell_car = generate_graph.find_total_income_by_selling_car(sheet_3,sheet_4)
sheet_5 = maintenance_func.refer_month(month,workbook_5)
total_maintenance_cost = generate_graph.find_total_income_by_maintenance(sheet_5)
total_income = total_income_sell_car + total_maintenance_cost
total_income_list = generate_graph.create_total_income_list(total_income)


month = 11
sheet = best_selling_models.refer_month(month,workbook)
max_row = sheet.max_row
model_list = best_selling_models.append_models_to_a_list(max_row,sheet)
model_dict = best_selling_models.make_a_dic_by_count_of_the_models(model_list)
print('\nBest five selling models for november 2021')
best_selling_models.print_best_five_selling_models_of_a_month(model_dict)
ask_price_list = monthly_profit.create_ask_price_list(sheet,max_row)
sell_price_list = monthly_profit.create_sell_price_list(sheet,max_row)
total_ask_price = monthly_profit.add_ask_prices(ask_price_list)
total_sell_price = monthly_profit.add_sell_prices(sell_price_list)
print('\nThe monthly profit generated by the sales department during november 2021')
monthly_profit_value = monthly_profit.print_monthly_profit(total_sell_price,total_ask_price)
sheet_3 = sold_details_append.refer_month(month,workbook_3)
sheet_4 = sold_details_append.refer_month_in_CompnyDetails_workbook(month,workbook_4)
customer_name = best_customers_or_companies.create_customer_list(sheet_3)
company_name = best_customers_or_companies.create_company_list(sheet_4)
customer_dict = best_customers_or_companies.make_a_dic_by_count_of_the_customers(customer_name)
company_dict = best_customers_or_companies.make_a_dic_by_count_of_the_companies(company_name)
print('\nAll best customers during november 2021')
best_customers_or_companies.find_best_customers(customer_dict)
print('\nAll best companies during november 2021')
best_customers_or_companies.find_best_companies(company_dict)
num_of_customers = sheet_3.max_row - 1
number_of_companies = sheet_4.max_row - 1
number_of_sold_cars = num_of_customers+number_of_companies
num_of_sold_car_list = generate_graph.create_a_list_with_number_of_sold_cars(number_of_sold_cars)
total_income_sell_car = generate_graph.find_total_income_by_selling_car(sheet_3,sheet_4)
sheet_5 = maintenance_func.refer_month(month,workbook_5)
total_maintenance_cost = generate_graph.find_total_income_by_maintenance(sheet_5)
total_income = total_income_sell_car + total_maintenance_cost
total_income_list = generate_graph.create_total_income_list(total_income)

month = 12
sheet = best_selling_models.refer_month(month,workbook)
max_row = sheet.max_row
model_list = best_selling_models.append_models_to_a_list(max_row,sheet)
model_dict = best_selling_models.make_a_dic_by_count_of_the_models(model_list)
print('\nBest five selling models for december 2021')
best_selling_models.print_best_five_selling_models_of_a_month(model_dict)
ask_price_list = monthly_profit.create_ask_price_list(sheet,max_row)
sell_price_list = monthly_profit.create_sell_price_list(sheet,max_row)
total_ask_price = monthly_profit.add_ask_prices(ask_price_list)
total_sell_price = monthly_profit.add_sell_prices(sell_price_list)
print('\nThe monthly profit generated by the sales department during december 2021')
monthly_profit_value = monthly_profit.print_monthly_profit(total_sell_price,total_ask_price)
sheet_3 = sold_details_append.refer_month(month,workbook_3)
sheet_4 = sold_details_append.refer_month_in_CompnyDetails_workbook(month,workbook_4)
customer_name = best_customers_or_companies.create_customer_list(sheet_3)
company_name = best_customers_or_companies.create_company_list(sheet_4)
customer_dict = best_customers_or_companies.make_a_dic_by_count_of_the_customers(customer_name)
company_dict = best_customers_or_companies.make_a_dic_by_count_of_the_companies(company_name)
print('\nAll best customers during december 2021')
best_customers_or_companies.find_best_customers(customer_dict)
print('\nAll best companies during december 2021')
best_customers_or_companies.find_best_companies(company_dict)
num_of_customers = sheet_3.max_row - 1
number_of_companies = sheet_4.max_row - 1
number_of_sold_cars = num_of_customers+number_of_companies
num_of_sold_car_list = generate_graph.create_a_list_with_number_of_sold_cars(number_of_sold_cars)
total_income_sell_car = generate_graph.find_total_income_by_selling_car(sheet_3,sheet_4)
sheet_5 = maintenance_func.refer_month(month,workbook_5)
total_maintenance_cost = generate_graph.find_total_income_by_maintenance(sheet_5)
total_income = total_income_sell_car + total_maintenance_cost
total_income_list = generate_graph.create_total_income_list(total_income)

month = 1
sheet = best_selling_models.refer_month(month,workbook)
max_row = sheet.max_row
model_list = best_selling_models.append_models_to_a_list(max_row,sheet)
model_dict = best_selling_models.make_a_dic_by_count_of_the_models(model_list)
print('\nBest five selling models for january 2022')
best_selling_models.print_best_five_selling_models_of_a_month(model_dict)
ask_price_list = monthly_profit.create_ask_price_list(sheet,max_row)
sell_price_list = monthly_profit.create_sell_price_list(sheet,max_row)
total_ask_price = monthly_profit.add_ask_prices(ask_price_list)
total_sell_price = monthly_profit.add_sell_prices(sell_price_list)
print('\nThe monthly profit generated by the sales department during january 2022')
monthly_profit_value = monthly_profit.print_monthly_profit(total_sell_price,total_ask_price)
sheet_3 = sold_details_append.refer_month(month,workbook_3)
sheet_4 = sold_details_append.refer_month_in_CompnyDetails_workbook(month,workbook_4)
customer_name = best_customers_or_companies.create_customer_list(sheet_3)
company_name = best_customers_or_companies.create_company_list(sheet_4)
customer_dict = best_customers_or_companies.make_a_dic_by_count_of_the_customers(customer_name)
company_dict = best_customers_or_companies.make_a_dic_by_count_of_the_companies(company_name)
print('\nAll best customers during january 2022')
best_customers_or_companies.find_best_customers(customer_dict)
print('\nAll best companies during january 2022')
best_customers_or_companies.find_best_companies(company_dict)
num_of_customers = sheet_3.max_row - 1
number_of_companies = sheet_4.max_row - 1
number_of_sold_cars = num_of_customers+number_of_companies
num_of_sold_car_list = generate_graph.create_a_list_with_number_of_sold_cars(number_of_sold_cars)
total_income_sell_car = generate_graph.find_total_income_by_selling_car(sheet_3,sheet_4)
sheet_5 = maintenance_func.refer_month(month,workbook_5)
total_maintenance_cost = generate_graph.find_total_income_by_maintenance(sheet_5)
total_income = total_income_sell_car + total_maintenance_cost
total_income_list = generate_graph.create_total_income_list(total_income)

month = 2
sheet = best_selling_models.refer_month(month,workbook)
max_row = sheet.max_row
model_list = best_selling_models.append_models_to_a_list(max_row,sheet)
model_dict = best_selling_models.make_a_dic_by_count_of_the_models(model_list)
print('\nBest five selling models for february 2022')
best_selling_models.print_best_five_selling_models_of_a_month(model_dict)
ask_price_list = monthly_profit.create_ask_price_list(sheet,max_row)
sell_price_list = monthly_profit.create_sell_price_list(sheet,max_row)
total_ask_price = monthly_profit.add_ask_prices(ask_price_list)
total_sell_price = monthly_profit.add_sell_prices(sell_price_list)
print('\nThe monthly profit generated by the sales department during february 2022')
monthly_profit_value = monthly_profit.print_monthly_profit(total_sell_price,total_ask_price)
sheet_3 = sold_details_append.refer_month(month,workbook_3)
sheet_4 = sold_details_append.refer_month_in_CompnyDetails_workbook(month,workbook_4)
customer_name = best_customers_or_companies.create_customer_list(sheet_3)
company_name = best_customers_or_companies.create_company_list(sheet_4)
customer_dict = best_customers_or_companies.make_a_dic_by_count_of_the_customers(customer_name)
company_dict = best_customers_or_companies.make_a_dic_by_count_of_the_companies(company_name)
print('\nAll best customers during february 2022')
best_customers_or_companies.find_best_customers(customer_dict)
print('\nAll best companies during february 2022')
best_customers_or_companies.find_best_companies(company_dict)
num_of_customers = sheet_3.max_row - 1
number_of_companies = sheet_4.max_row - 1
number_of_sold_cars = num_of_customers+number_of_companies
num_of_sold_car_list = generate_graph.create_a_list_with_number_of_sold_cars(number_of_sold_cars)
total_income_sell_car = generate_graph.find_total_income_by_selling_car(sheet_3,sheet_4)
sheet_5 = maintenance_func.refer_month(month,workbook_5)
total_maintenance_cost = generate_graph.find_total_income_by_maintenance(sheet_5)
total_income = total_income_sell_car + total_maintenance_cost
total_income_list = generate_graph.create_total_income_list(total_income)

# code to generate graphs
generate_graph.generate_graph_for_the_sold_cars(num_of_sold_car_list)
generate_graph.generate_graph_for_the_total_income(total_income_list)