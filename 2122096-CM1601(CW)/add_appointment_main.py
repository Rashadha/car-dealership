import openpyxl
import add_appointment_func
import update_record

# loading CarDealership.xlsx
workbook = openpyxl.load_workbook("CarDealership.xlsx")

# car models are only taken from '2022MARCH' sheet because I have used this sheet as current record sheet (it means I have taken the current month as march
sheet = workbook['2022MARCH']

# loading CarAppointment.xlsx
workbook_2 = openpyxl.load_workbook("CarAppointment.xlsx")

# current appointments which are belonging to each car uniquely will be added to this sheet
sheet_2 = workbook_2['appointment']

# declare varriables
option_1 = 1
number_list = []
number_list_2 = []
make_list = []
model_list = []
year_list = []
mileage_list = []
owner_info_list = []
ask_price_list = []
sell_price_list = []
car_status_list = []
last_price_status_list= []
last_price_list = []
new_list_2 = []
new_list_3 = []
car_record = []
appointment_list = []

# admin can add appointments until he selects the exit option
while option_1 == 1:
    max_row = sheet.max_row
    max_row_2 = sheet_2.max_row
    max_column = sheet_2.max_column
    # to view all the records of current month car details
    update_record.view_details(new_list_2, new_list_3, car_record, max_row, sheet, number_list, make_list,
                                      model_list, year_list,
                                      mileage_list, owner_info_list, ask_price_list, car_status_list, sell_price_list,
                               last_price_status_list,last_price_list)
    # to enter the record number to which car admin is going to add appointments
    detail_row_number = update_record.enter_record_number(max_row)

    # it will define 'is the admin entered record number belonging to the car details' and also 'isn't that make sold'
    for i in range(2, max_row + 1):
        if sheet.cell(i, 1).value == detail_row_number:
            if sheet.cell(i, 9).value == "not_sold":
                for j in range(2, max_row_2 + 1):
                    number_list_2.append(sheet_2.cell(j, 1).value)

    # if this condition true system will not enable you to add appointments bcs that you entered make was already sold
    if len(number_list_2) == 0:
        print('the car model was sold')

    # if this condition true that you entered make is not sold and also there is no any appointments belongs to that so you can add appointments to that
    elif detail_row_number not in number_list_2:
        add_appointment_func.append_appointments(sheet_2, max_row_2,max_column, detail_row_number, workbook_2,
                                                 appointment_list)
    # if this condition is true it will print below statements and also according to admin's preference it will allow admin to add more appointments to the earlier entered car
    elif detail_row_number in number_list_2:
        print('You have already entered appointments to that related car model')
        print('if you need to enter details to same model ---> press 1\nif not ---> press 2')
        option_2 = add_appointment_func.enter_second_option()
        if option_2 == 1:
            add_appointment_func.append_appointments(sheet_2, max_row_2,max_column, detail_row_number, workbook_2,
                                                     appointment_list)
    # to ask admin preference to add more appointments
    print('If you want to enter more records --> press 1\nif not --> press 2')
    option_1 = add_appointment_func.enter_first_option()
    appointment_list.clear()