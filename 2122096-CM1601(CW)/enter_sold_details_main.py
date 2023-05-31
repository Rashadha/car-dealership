import openpyxl

import input_car_details
import sold_details_append
import update_record

workbook = openpyxl.load_workbook("CarDealership.xlsx")

workbook_3 = openpyxl.load_workbook("CustomerInformation.xlsx")
workbook_4 = openpyxl.load_workbook("CompanyInformation.xlsx")

new_list_2 = []
new_list_3 = []
car_record = []
number_list = []
make_list = []
model_list = []
year_list = []
mileage_list = []
owner_info_list = []
ask_price_list = []
car_status_list = []
sell_price_list = []
last_price_status_list = []
last_price_list = []
name_list = []
address_list = []
contact_num_list = []
payment_type_list =[]
account_num_list = []
receipt_num_list = []
sold_price_list = []
option = 1


while option == 1:
    print('Enter the month that you need to add details')
    month = input_car_details.type_month()
    sheet = input_car_details.refer_month(month,workbook)
    max_row = sheet.max_row
    sheet_3 = sold_details_append.refer_month(month, workbook_3)
    max_row_3 = sheet_3.max_row
    sheet_4 = sold_details_append.refer_month_in_CompnyDetails_workbook(month,workbook_4)
    max_row_4 = sheet_4.max_row
    cus_list = sold_details_append.create_cus_list(max_row_3,sheet_3)
    comp_list = sold_details_append.create_comp_list(max_row_4,sheet_4)
    update_record.view_details(new_list_2, new_list_3, car_record, max_row, sheet, number_list, make_list, model_list, year_list,
                 mileage_list, owner_info_list, ask_price_list, car_status_list, sell_price_list,last_price_status_list,last_price_list)
    new_list_2.clear()
    new_list_3.clear()
    car_record.clear()
    number_list.clear()
    make_list.clear()
    model_list.clear()
    year_list.clear()
    mileage_list.clear()
    owner_info_list.clear()
    ask_price_list.clear()
    sell_price_list.clear()
    car_status_list.clear()
    last_price_status_list.clear()
    last_price_list.clear()

    detail_row_number = sold_details_append.enter_record_number_to_add_sold_info(max_row,cus_list,comp_list)
    print('If it\'s a customer ---> press 1\nIf it\'s a company ---> press 2')
    option_3 = update_record.enter_third_option()
    if option_3 == 1:
        customer_name = sold_details_append.enter_customer_name()
        customer_address = sold_details_append.enter_customer_address()
        customer_contact_num = sold_details_append.enter_customer_contact_number()
        payment_type = sold_details_append.enter_payment_type()
        account_number = sold_details_append.enter_account_number(payment_type)
        receipt_number = sold_details_append.enter_receipt_number()
        car_ref_num = detail_row_number
        car_make = sold_details_append.enter_car_make(max_row,sheet,detail_row_number)
        car_model = sold_details_append.enter_car_model(max_row,sheet,detail_row_number)
        car_year = sold_details_append.enter_car_year(max_row,sheet,detail_row_number)
        sold_price = sold_details_append.enter_sold_price(max_row,sheet,detail_row_number)

        name_list.append(customer_name)
        address_list.append(customer_address)
        contact_num_list.append(customer_contact_num)
        payment_type_list.append(payment_type)
        account_num_list.append(account_number)
        receipt_num_list.append(receipt_number)
        number_list.append(car_ref_num)
        make_list.append(car_make)
        model_list.append(car_model)
        year_list.append(car_year)
        sold_price_list.append(sold_price)

        sold_details_append.write_record_name_cus(name_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_address_cus(address_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_contact_num_cus(contact_num_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_payment_type_cus(payment_type_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_account_num_cus(account_num_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_receipt_num_cus(receipt_num_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_car_ref_num_cus(number_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_car_make_cus(make_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_car_model_cus(model_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_car_year_cus(year_list,sheet_3,max_row_3,workbook_3)
        sold_details_append.write_record_car_sold_price_cus(sold_price_list,sheet_3,max_row_3,workbook_3)
    else:
        company_name = sold_details_append.enter_company_name()
        company_location = sold_details_append.enter_company_location()
        company_contact_num = sold_details_append.enter_company_contact_number()
        payment_type = sold_details_append.enter_payment_type()
        account_number = sold_details_append.enter_account_number(payment_type)
        receipt_number = sold_details_append.enter_receipt_number()
        car_ref_num = detail_row_number
        car_make = sold_details_append.enter_car_make(max_row,sheet,detail_row_number)
        car_model = sold_details_append.enter_car_model(max_row,sheet,detail_row_number)
        car_year = sold_details_append.enter_car_year(max_row,sheet,detail_row_number)
        sold_price = sold_details_append.enter_sold_price(max_row, sheet, detail_row_number)

        name_list.append(company_name)
        address_list.append(company_location)
        contact_num_list.append(company_contact_num)
        payment_type_list.append(payment_type)
        account_num_list.append(account_number)
        receipt_num_list.append(receipt_number)
        number_list.append(car_ref_num)
        make_list.append(car_make)
        model_list.append(car_model)
        year_list.append(car_year)
        sold_price_list.append(sold_price)

        sold_details_append.write_record_name_comp(name_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_address_comp(address_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_contact_num_comp(contact_num_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_payment_type_comp(payment_type_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_account_num_comp(account_num_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_receipt_num_comp(receipt_num_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_car_ref_num_comp(number_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_car_make_comp(make_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_car_model_comp(model_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_car_year_comp(year_list, sheet_4, max_row_4, workbook_4)
        sold_details_append.write_record_car_sold_price_comp(sold_price_list, sheet_4, max_row_4, workbook_4)

    name_list.clear()
    address_list.clear()
    contact_num_list.clear()
    payment_type_list.clear()
    account_num_list.clear()
    receipt_num_list.clear()
    number_list.clear()
    make_list.clear()
    model_list.clear()
    year_list.clear()
    sold_price_list.clear()

    # if a car is sold to a customer, this will change that record car status column to sold on CarDealership.xlsx
    for i in range(2,max_row+1):
        if sheet.cell(i,1).value ==  detail_row_number:
            sheet.cell(row=i, column=9, value='sold')
            workbook.save('CarDealership.xlsx')

    print('If you want to add customer details or company details again ---> press 1\nif not ---> press 2')
    option = update_record.enter_option()
